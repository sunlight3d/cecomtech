import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

dir_path = r"C:\code\Cecomtech\HuynhTanKiet"
os.makedirs(dir_path, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "NhatKyHocTap"

# Define headers
headers = ["Ngày", "Nội dung học", "Giờ học", "Tổng số giờ", "Nhận xét"]
ws.append(headers)

# Styling
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12, name="Arial")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply styling to headers
for col_num, cell in enumerate(ws[1], 1):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Column widths
column_widths = {
    'A': 15,  # Ngày
    'B': 40,  # Nội dung học
    'C': 15,  # Giờ học
    'D': 15,  # Tổng số giờ
    'E': 40   # Nhận xét
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Add a few empty rows with text wrapping and border
wrap_alignment = Alignment(wrap_text=True, vertical="top")

for row in range(2, 30):
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.alignment = wrap_alignment
        cell.border = thin_border
        cell.font = Font(name="Arial", size=11)

# Save file
file_path = os.path.join(dir_path, "NhatKyHocTap.xlsx")
wb.save(file_path)
print(f"Excel file created successfully at {file_path}")
